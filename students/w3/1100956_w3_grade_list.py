#POP
import openpyxl
file="C:/Users/a0965/.spyder-py3/grade_list.xlsx"
work=openpyxl.load_workbook(file)
sheet1=work.sheetnames
sheet2=work[sheet1[0]]
range1=sheet2['c5:j12']
grade_list=[[data.value for data in row] for row in range1]
n=[]

for i in range(1,8):
    total = sum(grade_list[i][1:5])
    average = total / 4
    grade_list[i][5] = total
    grade_list[i][6] = average
    n.append(average)

n=sorted(n,reverse=True)
for i in range(1,8):
    grade_list[i][7]=n.index(grade_list[i][6]) + 1

for i in range(len(grade_list)):
    print(grade_list[i][0], end='\t')
    for j in range(1, len(grade_list[i])):
        print(grade_list[i][j], end='\t')
    print()
#OOP
import openpyxl

class Grade:
    def __init__(self, file_path):
        self.file = file_path
        self.workbook = openpyxl.load_workbook(self.file)
        self.sheet = self.workbook.active
        self.grade_list = []

    def extract_data(self):
        for row in self.sheet.iter_rows(min_row=5, max_row=12, min_col=3, max_col=10):
            row_data = [cell.value for cell in row]
            self.grade_list.append(row_data)

    def calculate(self):
        n = []
        for i in range(1,8):
            total = sum(float(value) if isinstance(value, (int, float)) else 0 for value in self.grade_list[i][1:5])
            average = total / 4
            self.grade_list[i][5] = total
            self.grade_list[i][6] = average
            n.append(average)

        n = sorted(n, reverse=True)
        for i in range(1,8):
            self.grade_list[i][7] = n.index(self.grade_list[i][6]) + 1

    def gradelist(self):
        for row in self.grade_list:
            for cell in row:
                print(cell, end='\t')
            print()

if __name__ == "__main__":
    file_path = "C:/Users/a0965/.spyder-py3/grade_list.xlsx"
    processor = Grade(file_path)
    processor.extract_data()
    processor.calculate()
    processor.gradelist()